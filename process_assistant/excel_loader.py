from __future__ import annotations

from pathlib import Path
from typing import Dict, List, Optional

from openpyxl import load_workbook

from .models import ProcessTime


def load_process_times_from_excel(excel_path: str | Path) -> List[ProcessTime]:
    wb = load_workbook(excel_path, data_only=True)
    records: List[ProcessTime] = []
    source_file = Path(excel_path).name

    for ws in wb.worksheets:
        sections = _extract_sections(ws, source_file)
        records.extend(sections)

    return records


def _extract_sections(ws, source_file: str) -> List[ProcessTime]:
    rows = [list(row) for row in ws.iter_rows(values_only=True)]
    records: List[ProcessTime] = []
    section_index = 0

    i = 0
    while i < len(rows) - 2:
        row = rows[i]
        if not _is_header_row(row):
            i += 1
            continue

        subheader = rows[i + 1]
        col_map = _build_column_map(row, subheader)
        section_index += 1
        section_name = f"{ws.title}_section_{section_index}"

        j = i + 2
        while j < len(rows):
            if _is_header_row(rows[j]):
                break

            process_name = _get(rows[j], col_map.get("process_name"))
            if process_name is None or str(process_name).strip() == "":
                # Empty rows in a section are tolerated.
                j += 1
                continue

            process_name = str(process_name).strip()
            if process_name in {"所属工段", "工序名称"}:
                j += 1
                continue

            record = ProcessTime(
                source_file=source_file,
                sheet_name=ws.title,
                section=section_name,
                process_id=_infer_process_id(process_name),
                process_name=process_name,
                sequence=_to_int(_get(rows[j], col_map.get("sequence"))),
                ct_sec=_to_float(_get(rows[j], col_map.get("ct_sec"))),
                defect_rate=_to_float(_get(rows[j], col_map.get("defect_rate"))),
                extra_factor=_to_float(_get(rows[j], col_map.get("extra_factor"))),
                ect_sec=_to_float(_get(rows[j], col_map.get("ect_sec"))),
                hourly_capacity=_to_float(_get(rows[j], col_map.get("hourly_capacity"))),
                takt_sec=_to_float(_get(rows[j], col_map.get("takt_sec"))),
                manpower=_to_float(_get(rows[j], col_map.get("manpower"))),
                daily_output=_to_float(_get(rows[j], col_map.get("daily_output"))),
                equipment=_to_str(_get(rows[j], col_map.get("equipment"))),
                equipment_ratio=_to_float(_get(rows[j], col_map.get("equipment_ratio"))),
                note=_to_str(_get(rows[j], col_map.get("note"))),
            )
            records.append(record)
            j += 1

        i = j

    return records


def _is_header_row(row: List[object]) -> bool:
    values = {str(x).strip() for x in row if x is not None and str(x).strip()}
    return "工序名称" in values and "所属工段" in values


def _build_column_map(header: List[object], subheader: List[object]) -> Dict[str, int]:
    m: Dict[str, int] = {}
    for idx, val in enumerate(header):
        text = str(val).strip() if val is not None else ""
        if text == "序号":
            m["sequence"] = idx
        elif text == "工序名称":
            m["process_name"] = idx
        elif "治工具" in text:
            m["equipment"] = idx
        elif text == "设备需求比":
            m["equipment_ratio"] = idx

    for idx, val in enumerate(subheader):
        text = str(val).strip() if val is not None else ""
        if text.startswith("C/T"):
            m["ct_sec"] = idx
        elif text == "不良率":
            m["defect_rate"] = idx
        elif "额外非生产时间" in text:
            m["extra_factor"] = idx
        elif text.startswith("E-SPCT"):
            m["ect_sec"] = idx
        elif "有效小时产能" in text:
            m["hourly_capacity"] = idx
        elif "工时节拍" in text:
            m["takt_sec"] = idx
        elif "日人力需求" in text:
            m["manpower"] = idx
        elif "实际日可产量" in text:
            m["daily_output"] = idx
        elif text == "备注":
            m["note"] = idx

    return m


def _infer_process_id(name: str) -> str:
    normalized = name.replace("\n", "")
    mapping = [
        (["预焊"], "P09"),
        (["焊线"], "P10"),
        (["模压"], "P11"),
        (["下料"], "P01"),
        (["印刷"], "P02"),
        (["烘烤", "固化"], "P03"),
        (["曝光"], "P04"),
        (["蚀刻", "显影"], "P05"),
        (["清洗", "清洁"], "P06"),
        (["贴合", "覆膜", "快压", "贴胶", "补强", "硅胶布"], "P07"),
        (["冲型", "冲孔", "冲定位"], "P08"),
        (["电阻", "耐压", "全检", "包装", "入库", "检查"], "P12"),
    ]
    for keywords, pid in mapping:
        if any(k in normalized for k in keywords):
            return pid
    return "P00"


def _get(row: List[object], idx: Optional[int]):
    if idx is None or idx < 0 or idx >= len(row):
        return None
    return row[idx]


def _to_float(value) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if text in {"", "N/A", "充足", "×1", "×2", "×4"}:
        return None
    text = text.replace(",", "")
    try:
        return float(text)
    except ValueError:
        return None


def _to_int(value) -> Optional[int]:
    if value is None:
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    text = str(value).strip().replace("*", "")
    if not text:
        return None
    try:
        return int(float(text))
    except ValueError:
        return None


def _to_str(value) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None
